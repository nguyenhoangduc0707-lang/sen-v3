import requests
import pandas as pd
import argparse
import time
import os
import zipfile
import pickle
import base64
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from datetime import datetime, timedelta
from typing import List, Dict, Optional
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Các thư viện Google API
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload
from googleapiclient.errors import HttpError

# Quyền truy cập Google API (Scopes)
SCOPES = [
    'https://www.googleapis.com/auth/gmail.send',
    'https://www.googleapis.com/auth/drive.file'  # Chỉ làm việc với các file do script này tạo ra
]

class AccessTradeFetcher:
    def __init__(self, api_key: str):
        # 1. Ép kiểu về chuỗi, loại bỏ khoảng trắng thừa hai đầu
        clean_key = str(api_key).strip()
        
        # 2. Loại bỏ triệt để mọi ký tự tiếng Việt có dấu hoặc ký tự lạ nếu vô tình dính vào
        clean_key = clean_key.encode('ascii', 'ignore').decode('ascii')
        
        self.api_key = clean_key
        self.base_url = "https://api.accesstrade.vn/v1/order-list"
        
        # 3. Đảm bảo toàn bộ tiêu đề Header gửi đi đạt chuẩn mã hóa Latin-1
        self.headers = {
            "Authorization": f"Bearer {clean_key}",
            "Content-Type": "application/json"
        }

    def fetch_orders(self, since: str, until: str, page: int = 1, limit: int = 300, 
                     status: Optional[int] = None, merchant: Optional[str] = None, 
                     utm_source: Optional[str] = None) -> Dict:
        """Gọi API lấy dữ liệu của một trang cụ thể"""
        params = {
            "since": since,
            "until": until,
            "page": page,
            "limit": min(limit, 300)
        }
        if status is not None:
            params["status"] = status
        if merchant:
            params["merchant"] = merchant
        if utm_source:
            params["utm_source"] = utm_source

        response = requests.get(self.base_url, headers=self.headers, params=params)
        if response.status_code == 200:
            return response.json()
        else:
            raise Exception(f"Lỗi API {response.status_code}: {response.text}")

    def fetch_all_orders(self, since: str, until: str, limit_per_page: int = 300, 
                         status: Optional[int] = None, merchant: Optional[str] = None, 
                         utm_source: Optional[str] = None) -> List[Dict]:
        """Tự động cuộn trang để lấy toàn bộ đơn hàng trong khoảng thời gian"""
        all_orders = []
        current_page = 1
        
        while True:
            print(f"📥 Đang lấy dữ liệu trang {current_page}...")
            # Xử lý Rate Limit: Tối đa 10 requests/phút -> Nghỉ 6 giây giữa các lần gọi
            if current_page > 1:
                time.sleep(6)
                
            result = self.fetch_orders(
                since=since, until=until, page=current_page, 
                limit=limit_per_page, status=status, 
                merchant=merchant, utm_source=utm_source
            )
            orders = result.get("data", [])
            total = result.get("total", 0)
            
            if not orders:
                break
                
            all_orders.extend(orders)
            print(f"   ✅ Đã tải: {len(all_orders)} / {total} đơn hàng")
            
            if len(all_orders) >= total:
                break
            current_page += 1
            
        return all_orders

    def process_data(self, orders: List[Dict]) -> pd.DataFrame:
        """Xử lý thô dữ liệu JSON chuyển đổi thành DataFrame"""
        if not orders:
            return pd.DataFrame()
            
        df = pd.DataFrame(orders)
        columns_order = [
            'order_id', 'merchant', 'billing', 'pub_commission', 'is_confirmed', 
            'status', 'sales_time', 'confirmed_time', 'click_time', 'utm_source', 
            'utm_medium', 'utm_campaign', 'utm_content', 'client_platform', 'browser', 
            'products_count', 'category_name', 'product_category'
        ]
        existing_columns = [col for col in columns_order if col in df.columns]
        df = df[existing_columns]
        
        status_map = {0: 'Pending/Hold', 1: 'Approved', 2: 'Rejected'}
        df['status'] = df['is_confirmed'].map(status_map)
        
        if 'sales_time' in df.columns:
            df['sales_time'] = pd.to_datetime(df['sales_time'])
        if 'confirmed_time' in df.columns:
            df['confirmed_time'] = pd.to_datetime(df['confirmed_time'])
            
        return df

    def optimize_excel_size(self, df: pd.DataFrame, filename: str, optimize: bool = False):
        """Xuất dữ liệu ra file Excel định dạng chuẩn và đẹp mắt"""
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Chi tiết đơn hàng', index=False)
            
            if not optimize:
                # Tạo thêm sheet thống kê theo nhà bán hàng (Merchant)
                if 'merchant' in df.columns and 'pub_commission' in df.columns:
                    merchant_stats = df.groupby('merchant').agg({
                        'order_id': 'count', 'billing': 'sum', 'pub_commission': 'sum'
                    }).round(2).reset_index()
                    merchant_stats.columns = ['Merchant', 'Số đơn', 'Tổng giá trị (VNĐ)', 'Tổng hoa hồng (VNĐ)']
                    merchant_stats.to_excel(writer, sheet_name='Thống kê Merchant', index=False)
                
                # Tạo thêm sheet thống kê theo chiến dịch UTM Source
                if 'utm_source' in df.columns and df['utm_source'].notna().any():
                    utm_stats = df[df['utm_source'].notna()].groupby('utm_source').agg({
                        'order_id': 'count', 'pub_commission': 'sum'
                    }).round(2).reset_index()
                    utm_stats.columns = ['UTM Source', 'Số đơn', 'Tổng hoa hồng (VNĐ)']
                    utm_stats.to_excel(writer, sheet_name='Thống kê UTM', index=False)
                    
        # Tự động căn chỉnh độ rộng cột Excel
        wb = load_workbook(filename)
        for sheet in wb.worksheets:
            for column in sheet.columns:
                max_len = max(len(str(cell.value or '')) for cell in column)
                col_letter = column[0].column_letter
                sheet.column_dimensions[col_letter].width = min(max_len + 3, 30)
                
            # Đổ màu Header cho chuyên nghiệp
            for cell in sheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
        wb.save(filename)
        print(f"✅ Đã xuất báo cáo Excel: {filename}")
        return os.path.getsize(filename) / (1024 * 1024)

    def export_csv_compressed(self, df: pd.DataFrame, filename: str):
        """Xuất file CSV mã hóa UTF-8 và nén thành file định dạng .zip"""
        csv_filename = filename.replace('.zip', '.csv')
        df.to_csv(csv_filename, index=False, encoding='utf-8-sig')
        
        with zipfile.ZipFile(filename, 'w', zipfile.ZIP_DEFLATED) as zipf:
            zipf.write(csv_filename, arcname=os.path.basename(csv_filename))
        os.remove(csv_filename)
        print(f"✅ Đã xuất file nén CSV ZIP: {filename}")


class GoogleServices:
    def __init__(self, credentials_file='credentials.json', token_file='token.pickle'):
        self.credentials_file = credentials_file
        self.token_file = token_file
        self.gmail_service = None
        self.drive_service = None
        self._authenticate()

    def _authenticate(self):
        """Xác thực thực thể bằng OAuth 2.0 (mở browser trong lần đầu tiên chạy)"""
        creds = None
        if os.path.exists(self.token_file):
            with open(self.token_file, 'rb') as token:
                creds = pickle.load(token)
                
        if not creds or not creds.valid:
            if creds and creds.expired and creds.refresh_token:
                creds.refresh(Request())
            else:
                if not os.path.exists(self.credentials_file):
                    raise Exception(f"❌ Không tìm thấy file {self.credentials_file} trong thư mục hiện tại!")
                flow = InstalledAppFlow.from_client_secrets_file(self.credentials_file, SCOPES)
                creds = flow.run_local_server(port=0)
            with open(self.token_file, 'wb') as token:
                pickle.dump(creds, token)
                
        self.gmail_service = build('gmail', 'v1', credentials=creds)
        self.drive_service = build('drive', 'v3', credentials=creds)
        print("🔐 Đã đồng bộ cấu hình tài khoản Google thành công!")

    def send_email_gmail_api(self, to_email: str, subject: str, body_html: str, attachments: List[str] = None) -> bool:
        """Gửi email HTML bảo mật thông qua Gmail API"""
        try:
            message = MIMEMultipart('alternative')
            message['To'] = to_email
            message['From'] = os.environ.get('GMAIL_SENDER')
            message['Subject'] = subject
            
            message.attach(MIMEText(body_html, 'html'))
            
            if attachments:
                for file_path in attachments:
                    if os.path.exists(file_path):
                        with open(file_path, 'rb') as f:
                            mime_part = MIMEBase('application', 'octet-stream')
                            mime_part.set_payload(f.read())
                        encoders.encode_base64(mime_part)
                        mime_part.add_header('Content-Disposition', f'attachment; filename="{os.path.basename(file_path)}"')
                        message.attach(mime_part)
                        
            raw_message = base64.urlsafe_b64encode(message.as_bytes()).decode('utf-8')
            self.gmail_service.users().messages().send(userId='me', body={'raw': raw_message}).execute()
            print(f"📧 Báo cáo đã gửi thành công tới hòm thư: {to_email}")
            return True
        except Exception as e:
            print(f"❌ Gặp lỗi hệ thống khi gửi Email: {str(e)}")
            return False

    def upload_to_drive(self, file_path: str, folder_id: str = None, public_read: bool = True) -> Optional[str]:
        """Tự động tải tài liệu báo cáo lên Google Drive và xuất liên kết chia sẻ"""
        try:
            file_name = os.path.basename(file_path)
            file_metadata = {'name': file_name}
            if folder_id:
                file_metadata['parents'] = [folder_id]
                
            media = MediaFileUpload(file_path, resumable=True)
            file = self.drive_service.files().create(body=file_metadata, media_body=media, fields='id, webViewLink').execute()
            
            if public_read:
                permission = {'type': 'anyone', 'role': 'reader'}
                self.drive_service.permissions().create(fileId=file['id'], body=permission).execute()
                
            print(f"☁️ Đã đẩy file lên đám mây lưu trữ Google Drive: {file_name}")
            return file.get('webViewLink')
        except HttpError as error:
            print(f"❌ Gặp lỗi trong quá trình đồng bộ Drive: {error}")
            return None


def generate_email_body(df: pd.DataFrame, start_date: str, end_date: str, days: int, status: int, drive_links: Dict = None) -> str:
    """Tạo giao diện Email trực quan định dạng HTML"""
    total_orders = len(df)
    total_billing = df['billing'].sum() if 'billing' in df.columns else 0
    total_commission = df['pub_commission'].sum() if 'pub_commission' in df.columns else 0
    
    drive_section = ""
    if drive_links:
        drive_section = "<h2>📁 Đường dẫn tải file trực tiếp trên Google Drive:</h2>"
        for f_type, link in drive_links.items():
            drive_section += f'<p>👉 Định dạng {f_type.upper()}: <a href="{link}">Nhấn vào đây để xem dữ liệu</a></p>'
            
    status_text = {0: "Pending/Hold", 1: "Approved", 2: "Rejected"}
    
    html = f"""
    <html>
    <head>
        <style>
            body {{ font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; color: #333; }}
            .card {{ background: #f8f9fa; padding: 20px; border-radius: 8px; border: 1px solid #e3e6f0; }}
            .highlight {{ font-size: 20px; font-weight: bold; color: #4e73df; }}
        </style>
    </head>
    <body>
        <div class="card">
            <h2>📊 BÁO CÁO KẾT QUẢ DOANH THU ACCESSTRADE</h2>
            <p>🗓️ <b>Khoảng thời gian:</b> {start_date} &rarr; {end_date} ({days} ngày)</p>
            <p>⚙️ <b>Trạng thái đơn:</b> {status_text.get(status, 'Tất cả')}</p>
            <hr>
            <p>📦 <b>Tổng số lượng đơn phát sinh:</b> <span class="highlight">{total_orders:,}</span> đơn</p>
            <p>💰 <b>Tổng giá trị giao dịch:</b> <span class="highlight">{total_billing:,.0f} VNĐ</span></p>
            <p>💵 <b>Tổng mức hoa hồng nhận về:</b> <span class="highlight" style="color:#1cc88a;">{total_commission:,.0f} VNĐ</span></p>
            {drive_section}
        </div>
    </body>
    </html>
    """
    return html


def main():
    parser = argparse.ArgumentParser(description='AccessTrade Analytics Engine CLI')
    parser.add_argument('--days', type=int, default=30)
    parser.add_argument('--status', type=int, choices=[0, 1, 2], default=1)
    parser.add_argument('--merchant', type=str)
    parser.add_argument('--utm_source', type=str)
    parser.add_argument('--output', type=str, default='accesstrade_report')
    parser.add_argument('--format', choices=['excel', 'csv_zip', 'both'], default='both')
    parser.add_argument('--send-email', action='store_true')
    parser.add_argument('--email-to', type=str)
    parser.add_argument('--upload-drive', action='store_true')
    parser.add_argument('--drive-folder-id', type=str)
    parser.add_argument('--drive-upload-threshold-mb', type=int, default=20)
    parser.add_argument('--optimize-excel', action='store_true')
    args = parser.parse_args()

    api_key = os.environ.get('ACCESS_TRADE_API_KEY')
    if not api_key:
        raise Exception("❌ Lỗi: Hãy thiết lập biến môi trường cấu hình: ACCESS_TRADE_API_KEY")
        
    if args.send_email and not os.environ.get('GMAIL_SENDER'):
        raise Exception("❌ Lỗi: Hãy thiết lập biến môi trường: GMAIL_SENDER")

    end_date = datetime.now()
    start_date = end_date - timedelta(days=args.days)
    since = start_date.strftime("%Y-%m-%dT%H:%M:%SZ")
    until = end_date.strftime("%Y-%m-%dT%H:%M:%SZ")

    print(f"🚀 Khởi chạy trình trích xuất AccessTrade API...")
    fetcher = AccessTradeFetcher(api_key)
    orders = fetcher.fetch_all_orders(since=since, until=until, status=args.status, 
                                      merchant=args.merchant, utm_source=args.utm_source)
    
    if not orders:
        print("⚠️ Không tìm thấy dữ liệu đơn hàng nào khớp với yêu cầu tìm kiếm.")
        return

    df = fetcher.process_data(orders)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    attachments = []
    drive_links = {}
    need_upload = False

    if args.format in ['excel', 'both']:
        excel_file = f"{args.output}_{timestamp}.xlsx"
        size_mb = fetcher.optimize_excel_size(df, excel_file, args.optimize_excel)
        attachments.append(excel_file)
        if args.upload_drive and size_mb > args.drive_upload_threshold_mb:
            need_upload = True

    if args.format in ['csv_zip', 'both']:
        zip_file = f"{args.output}_{timestamp}.zip"
        fetcher.export_csv_compressed(df, zip_file)
        attachments.append(zip_file)
        zip_size = os.path.getsize(zip_file) / (1024 * 1024)
        if args.upload_drive and zip_size > args.drive_upload_threshold_mb:
            need_upload = True

    google = None
    if args.send_email or (args.upload_drive and need_upload):
        google = GoogleServices()

    if args.upload_drive and need_upload:
        for file_path in attachments:
            link = google.upload_to_drive(file_path, args.drive_folder_id)
            if link:
                f_type = 'excel' if file_path.endswith('.xlsx') else 'zip'
                drive_links[f_type] = link

    if args.send_email and google:
        subject = f"📊 Báo cáo tự động AccessTrade ngày {datetime.now().strftime('%Y-%m-%d')}"
        body = generate_email_body(df, since, until, args.days, args.status, drive_links)
        final_attachments = [] if need_upload else attachments
        google.send_email_gmail_api(args.email_to, subject, body, final_attachments)

    print("✨ Toàn bộ quy trình đã xử lý hoàn tất hoàn hảo!")

if __name__ == "__main__":
    main()